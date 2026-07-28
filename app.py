import streamlit as st
import openpyxl
import pandas as pd
import io
from datetime import datetime

# Page Configuration
st.set_page_config(
    page_title="Supplier Evaluation Portal",
    page_icon="📋",
    layout="wide"
)

st.title("📋 Supplier Evaluation Form Auto-Processor")
st.write("Upload the Sample Data Sheet and Supplier Evaluation Forms. The system will automatically parse and append all form data into the sample data sheet.")

st.divider()

# File Uploaders
col1, col2 = st.columns(2)

with col1:
    st.subheader("1️⃣ Sample Data Sheet")
    master_file = st.file_uploader("Upload Sample Data Sheet", type=["xlsx"], key="master")

with col2:
    st.subheader("2️⃣ Supplier Evaluation Forms")
    forms_files = st.file_uploader("Upload Supplier Evaluation Forms (Single or Multiple)", type=["xlsx"], accept_multiple_files=True, key="forms")

st.divider()

def parse_form_data(ws_form, next_id):
    def get_val(r, c):
        v = ws_form.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    def get_first_non_empty(r, cols):
        """Scans multiple columns in a row to find the first non-empty value."""
        for c in cols:
            v = get_val(r, c)
            if v:
                return v
        return ""

    def parse_rating_smart(row, col_pairs, default_keywords=None):
        """
        Robust rating parser:
        1. Direct check: Pairs of (num_col, text_col)
        2. Row Scan: Scans entire row for rating patterns/keywords (1, 2, 3, Frequent, Some, Always)
        """
        # Method 1: Check specific column pairs
        for num_col, text_col in col_pairs:
            val_num = get_val(row, num_col)
            val_text = get_val(row, text_col)
            
            # Direct numeric match
            if val_num in ['1', '2', '3']:
                if val_text:
                    return f"{val_num}-{val_text}"
                elif val_num == '1': return "1-Frequent delays"
                elif val_num == '2': return "2-Some delays"
                elif val_num == '3': return "3-Always on time"

        # Method 2: Scan broad column range in the row for checkboxes/numbers/labels
        row_text = ""
        found_num = None
        for col in range(1, 25):  # Scan columns A to X
            val = get_val(row, col)
            if val in ['1', '2', '3']:
                found_num = val
            if val:
                row_text += " " + val.lower()

        # Map according to detected values/keywords
        if 'frequent' in row_text or 'frequent delays' in row_text or found_num == '1':
            return "1-Frequent delays"
        elif 'some' in row_text or 'some delays' in row_text or found_num == '2':
            return "2-Some delays"
        elif 'always' in row_text or 'always on time' in row_text or found_num == '3':
            return "3-Always on time"

        return ""

    # Smart fallback for Evaluation Date
    raw_date = get_first_non_empty(12, [6, 5, 7, 8, 14])
    eval_date = raw_date.split(' ')[0] if raw_date else datetime.now().strftime('%Y-%m-%d')
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Common search column ranges to prevent missing data if vendor shifted cells
    col_search_main = [6, 5, 7, 8, 9, 10, 14]
    col_search_contact = [14, 13, 15, 6, 8]

    vendor_name = get_first_non_empty(6, col_search_main)
    contact_person = get_first_non_empty(6, col_search_contact) or vendor_name

    # Universal column pair checks
    rating_cols = [(13, 14), (10, 11), (11, 12), (12, 13), (15, 16), (17, 18), (19, 20)]

    return {
        'Id': next_id,
        'Start time': now_str,
        'Completion time': now_str,
        'Email': "muhammad.saad1@tcf.org.pk",
        'Name': "Muhammad Saad",
        'Company Name:': vendor_name,
        'Location / Address:': get_first_non_empty(8, col_search_main),
        'Supplier ID:': get_first_non_empty(10, col_search_main),
        'Evaluation Date:': eval_date,
        'Contact Person': contact_person,
        'Contact Number:': get_first_non_empty(8, col_search_contact),
        'Product/Service Provided:': get_first_non_empty(10, col_search_contact),
        'Period Evaluated:1': get_first_non_empty(12, col_search_contact),
        'Evaluation Cycle:': 'Annually',
        'In Business for:': 'More than 5 Years',
        'Registered vendor at TCF for:': 'More than 5 Years',
        'Quality Standard Compliance:': 'Others',

        # Ratings, Measures & Remarks
        'Quality of Product / Service (Weightage%)': get_first_non_empty(29, [8, 7, 9, 6]),
        'Quality of Product / Service: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(29, rating_cols),
        'Quality of Product / Service:\xa0Measure (If applicable)': get_first_non_empty(31, [8, 7, 9, 10]),
        'Quality of Product / Service (Remarks)': get_first_non_empty(33, [8, 7, 9, 10]),

        'Consistency of Quality (Weightage%)': get_first_non_empty(36, [8, 7, 9, 6]),
        'Consistency of Quality: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(36, rating_cols),
        'Consistency of Quality:\xa0Measure (If applicable)': get_first_non_empty(38, [8, 7, 9, 10]),
        'Consistency of Quality (Remarks)': get_first_non_empty(39, [8, 7, 9]) or get_first_non_empty(40, [8, 7, 9]),

        'Meets Specifications (Weightage%)': get_first_non_empty(43, [8, 7, 9, 6]),
        'Meets Specifications: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(43, rating_cols),
        'Meets Specifications:\xa0Measure (If applicable)': get_first_non_empty(45, [8, 7, 9, 10]),
        'Meets Specifications (Remarks)': get_first_non_empty(47, [8, 7, 9, 10]),

        'Pricing compared to market rates (Weightage%)': get_first_non_empty(55, [8, 7, 9, 6]),
        'Pricing compared to market rates: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(55, rating_cols),
        'Pricing compared to market rates:\xa0Measure (If applicable)': get_first_non_empty(57, [8, 7, 9, 10]),
        'Pricing compared to market rates (Remarks)': get_first_non_empty(59, [8, 7, 9, 10]),

        'Value for Money (Weightage%)': get_first_non_empty(62, [8, 7, 9, 6]),
        'Value for Money: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(62, rating_cols),
        'Value for Money:\xa0Measure (If applicable)': get_first_non_empty(64, [8, 7, 9, 10]),
        'Value for Money (Remarks)': get_first_non_empty(66, [8, 7, 9, 10]),

        'Discounts/Negotiations Provided (Weightage%)': get_first_non_empty(69, [8, 7, 9, 6]),
        'Discounts/Negotiations Provided: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(69, rating_cols),
        'Discounts/Negotiations Provided:\xa0Measure (If applicable)': get_first_non_empty(71, [11, 8, 7, 9, 10]),
        'Discounts/Negotiations Provided (Remarks)': get_first_non_empty(73, [8, 7, 9, 10]),

        'Timely Delivery (Weightage%)': get_first_non_empty(81, [8, 7, 9, 6]),
        'Timely Delivery: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(81, rating_cols),
        'Timely Delivery:\xa0Measure (If applicable)': get_first_non_empty(83, [8, 7, 9, 10]),
        'Timely Delivery (Remarks)': get_first_non_empty(85, [8, 7, 9, 10]),

        'Condition of Product upon delivery (Weightage%)': get_first_non_empty(88, [8, 7, 9, 6]),
        'Condition of Product upon delivery: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(88, rating_cols),
        'Condition of Product upon delivery:\xa0Measure (If applicable)': get_first_non_empty(90, [10, 8, 7, 9]),
        'Condition of Product upon delivery (Remarks)': get_first_non_empty(91, [8, 7, 9]) or get_first_non_empty(92, [8, 7, 9]),

        'Flexibility in Urgent requirements (Weightage%)': get_first_non_empty(95, [8, 7, 9, 6]),
        'Flexibility in Urgent requirements: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(95, rating_cols),
        'Flexibility in Urgent requirements:\xa0Measure (If applicable)': get_first_non_empty(97, [10, 8, 7, 9]),
        'Flexibility in Urgent requirements: (Remarks)': get_first_non_empty(99, [8, 7, 9, 10]),

        'Responsiveness to Queries/Concerns (Weightage%)': get_first_non_empty(108, [8, 7, 9, 6]),
        'Responsiveness to Queries/Concerns: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(108, rating_cols),
        'Responsiveness to Queries/Concerns:\xa0Measure (If applicable)': get_first_non_empty(110, [8, 7, 9, 10]),
        'Responsiveness to Queries/Concerns:\xa0(Remarks)': get_first_non_empty(112, [8, 7, 9, 10]),

        'After-Sales Support/Services (Weightage%)': get_first_non_empty(115, [8, 7, 9, 6]),
        'After-Sales Support/Services: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(115, rating_cols),
        'After-Sales Support/Services:\xa0Measure (If applicable)': get_first_non_empty(117, [10, 8, 7, 9]),
        'After-Sales Support/Services: (Remarks)': get_first_non_empty(119, [8, 7, 9, 10]),

        'Professionalism (Weightage%)': get_first_non_empty(122, [8, 7, 9, 6]),
        'Professionalism: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(122, rating_cols),
        'Professionalism:\xa0Measure (If applicable)': get_first_non_empty(124, [8, 7, 9, 10]),
        'Professionalism: (Remarks)': get_first_non_empty(126, [8, 7, 9, 10]),

        'Adhere to Legal & Ethical standards (Weightage%)': get_first_non_empty(134, [8, 7, 9, 6]),
        'Adhere to Legal & Ethical standards: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(134, rating_cols),
        'Adhere to Legal & Ethical standards:\xa0Measure (If applicable)': get_first_non_empty(136, [10, 8, 7, 9]),
        'Adhere to Legal & Ethical standards:\xa0(Remarks)': get_first_non_empty(137, [8, 7, 9]) or get_first_non_empty(138, [8, 7, 9]),

        'Accuracy of invoices & Documents (Weightage%)': get_first_non_empty(140, [8, 7, 9, 6]),
        'Accuracy of invoices & Documents: High = 3 (>95%),\xa0Medium = 2 (>80%),\xa0Low = 1 (<80%)': parse_rating_smart(140, rating_cols),
        'Accuracy of invoices & Documents: Measure (If applicable)': get_first_non_empty(142, [8, 7, 9, 10]),
        'Accuracy of invoices & Documents: (Remarks)': get_first_non_empty(144, [8, 7, 9, 10]),

        'Overall Comments:': get_first_non_empty(170, [5, 6, 7, 8, 4])
    }

# Process Action
if st.button("🚀 Merge & Update Sample Data Sheet", type="primary", use_container_width=True):
    if not master_file:
        st.error("⚠️ Please upload the Sample Data Sheet first!")
    elif not forms_files:
        st.error("⚠️ Please upload at least one Supplier Evaluation Form!")
    else:
        with st.spinner("Processing forms and merging data..."):
            dest_wb = openpyxl.load_workbook(io.BytesIO(master_file.getvalue()))
            dest_ws = dest_wb.active
            headers = [dest_ws.cell(row=1, column=c).value for c in range(1, dest_ws.max_column + 1)]

            count = 0
            for form_file in forms_files:
                form_wb = openpyxl.load_workbook(io.BytesIO(form_file.getvalue()), data_only=True)
                
                # Check for LightupEnterprises sheet first, otherwise fall back to the active worksheet
                if 'LightupEnterprises' in form_wb.sheetnames:
                    ws_form = form_wb['LightupEnterprises']
                else:
                    ws_form = form_wb.active

                next_id = dest_ws.max_row
                row_dict = parse_form_data(ws_form, next_id)
                new_row = [row_dict.get(h, "") for h in headers]
                dest_ws.append(new_row)
                count += 1

            output_stream = io.BytesIO()
            dest_wb.save(output_stream)
            output_stream.seek(0)

            st.success(f"🎉 Success! {count} supplier form(s) merged into the sample data sheet successfully.")

            st.download_button(
                label="📥 Download Updated Sample Data Sheet",
                data=output_stream,
                file_name="Sample_Data_Sheet_Updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
